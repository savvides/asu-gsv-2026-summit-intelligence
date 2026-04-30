#!/usr/bin/env python3
"""Match transcript files to schedule entries via fuzzy title overlap.

Reads:
  - private/source-data/ASU_GSV_2026_Schedule.xlsx (Full Schedule sheet)
  - repo/transcripts/*.md (transcript filenames give us NN-slug)
  - private/source-audio/NN - Title.mp3 (gives us full yt_title)
  - repo/video-metadata.json (existing entries, to skip)

Produces:
  - Proposed video-metadata.json entries for any transcript not yet in the JSON.
  - Surfaces low-confidence matches for human review.
"""

import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
ROOT = REPO.parent
XLSX = ROOT / "private" / "source-data" / "ASU_GSV_2026_Schedule.xlsx"
TRANSCRIPTS = REPO / "transcripts"
SOURCE_AUDIO = ROOT / "private" / "source-audio"
META_JSON = REPO / "video-metadata.json"

# Playlist YouTube IDs for the new videos (gathered from yt-dlp earlier)
YT_IDS = {
    "48": "Q5BmK8GY8OU", "49": "1YEuFjIZ-4Y", "50": "5ATt1FQWLWo",
    "52": "PI_tQE4_yHM", "54": "MQG7JZ4KjeA", "57": "CrxtL_VV6_k",
    "59": "AwUNXh822qo", "60": "v9yxays1X6o", "61": "2KRMYmDUax0",
    "62": "GPeMohBI0Jc", "63": "QxJvXKHez4A", "64": "6WQpwAl8Tm0",
    "65": "VDkAp8rzCjQ", "67": "vk5OQp90Xp0", "71": "O1O8dzLERTQ",
    "72": "OcV9LLxikCA", "73": "MZXZt-wlsF0", "74": "ToPR6aeT9gQ",
    "77": "KFbKgI8KUfE", "78": "fYD5htGmozA", "79": "OmiENmsh3dY",
    "80": "Rz5tFWsdkhA", "82": "n66PXC1E5Gw", "84": "pIti7HVYky0",
    "85": "CYVmoFw83R8", "87": "502SY2xODgw", "88": "fkuxV0H2Lvg",
    "91": "YQXLcoD75E4", "93": "TdCW4xqirU4", "94": "Ll89BW6CBuk",
}

STOP = set("a an the of and to in on for with or but is are was were be been being "
           "this that these those it its at as by from how what why when where who".split())

# Manual overrides where automatic matching can't find the right schedule row.
# Maps transcript num -> schedule Title (must match exactly). None means "no schedule match;
# leave schedule fields blank but keep the entry."
MANUAL = {
    "59": None,  # "Paradox of Perfect Machines" — not in schedule
    "61": "StageX Tuesday Evening & Education Innovation Showcase",
    "67": "StageX Tuesday Breakfast",  # Selingo/Belkin conversation
    "73": None,  # "GSV Cup Fan Favorites" — closest match is one of three GSV Cup 50 panels
    "82": None,  # "Beyond the Pilot" — not in schedule
    "91": "StageX Wednesday Lunch",  # 278M Reasons / Andrew Ng + Greg Hart
    "93": "StageX Monday Lunch",  # Profit Prophet / George Gilder
}


def tokens(s: str) -> set[str]:
    s = re.sub(r"[^a-z0-9 ]", " ", (s or "").lower())
    return {t for t in s.split() if t and t not in STOP and len(t) > 1}


def jaccard(a: set[str], b: set[str]) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def descr_score(yt_title: str, descr: str) -> float:
    """Heuristic: how well does the yt_title appear inside the description?
    StageX block descriptions list the panels within them, so a high overlap
    of yt_title's distinctive phrases inside the description is a strong signal.
    """
    if not descr:
        return 0.0
    yt_norm = re.sub(r"[^a-z0-9 ]", " ", yt_title.lower())
    descr_norm = re.sub(r"[^a-z0-9 ]", " ", descr.lower())
    yt_tok = [t for t in yt_norm.split() if t and t not in STOP and len(t) > 2]
    if not yt_tok:
        return 0.0
    hits = sum(1 for t in yt_tok if t in descr_norm)
    return hits / len(yt_tok)


def load_schedule() -> list[dict]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Full Schedule"]
    rows = list(ws.iter_rows(values_only=True))
    header = [h for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        d = dict(zip(header, r))
        if not d.get("URL"):
            continue
        out.append(d)
    return out


def yt_title_for(num: str) -> str | None:
    """Read the actual yt title from the source-audio mp3 filename."""
    pattern = f"{num} - "
    for f in SOURCE_AUDIO.glob(f"{num} - *.mp3"):
        name = f.stem
        # Strip leading "NN - " and trailing " ｜ ASU+GSV Summit YYYY"
        title = re.sub(r"^\d+\s*-\s*", "", name)
        title = re.sub(r"\s*[｜|]\s*ASU\+GSV Summit \d{4}\s*$", "", title)
        # Filename uses fullwidth ： for : — restore
        title = title.replace("：", ":")
        return title.strip()
    return None


def transcript_for(num: str) -> str | None:
    for f in TRANSCRIPTS.glob(f"{num}-*.md"):
        return f.name
    return None


def main():
    schedule = load_schedule()
    print(f"Loaded {len(schedule)} schedule entries", file=sys.stderr)

    with open(META_JSON) as f:
        existing = json.load(f)
    existing_nums = {e["num"] for e in existing}
    print(f"Existing metadata: {len(existing)} entries; nums: {sorted(existing_nums)[:5]}...", file=sys.stderr)

    new_entries = []
    for num in sorted(YT_IDS.keys(), key=int):
        if num in existing_nums:
            continue
        yt_title = yt_title_for(num)
        if not yt_title:
            print(f"SKIP {num}: no source-audio file", file=sys.stderr)
            continue
        tfile = transcript_for(num)
        if not tfile:
            print(f"SKIP {num}: no transcript file", file=sys.stderr)
            continue

        yt_tok = tokens(yt_title)
        # Score every schedule row by max(title_jaccard, 0.85 * descr_overlap)
        scored = []
        for s in schedule:
            title_sc = jaccard(yt_tok, tokens(s["Title"]))
            d_sc = descr_score(yt_title, s.get("Description") or "") * 0.85
            sc = max(title_sc, d_sc)
            scored.append((sc, title_sc, d_sc, s))
        scored.sort(key=lambda x: -x[0])
        best_score, best_title_sc, best_d_sc, best = scored[0]

        # Apply manual override if specified
        if num in MANUAL:
            target = MANUAL[num]
            if target is None:
                best = None
                best_score = -1.0
            else:
                match = next((s for s in schedule if s["Title"] == target), None)
                if match is None:
                    print(f"  WARN: manual override target not found: {target!r}", file=sys.stderr)
                else:
                    best = match
                    best_score = 1.0  # marker for manual
                    best_title_sc = best_d_sc = 1.0

        if best is None:
            entry = {
                "num": num,
                "yt_title": yt_title,
                "yt_url": f"https://www.youtube.com/watch?v={YT_IDS[num]}",
                "speakers": "",
                "match_score": 0.0,
                "schedule_title": "",
                "track": "Unknown",
                "tags": "",
                "day": "",
                "time": "",
                "format": "",
                "schedule_url": "",
                "schedule_description": "",
                "transcript_file": tfile,
            }
            print(f"  {num} [NONE] {yt_title[:55]} — no schedule match (manual)", file=sys.stderr)
        else:
            entry = {
                "num": num,
                "yt_title": yt_title,
                "yt_url": f"https://www.youtube.com/watch?v={YT_IDS[num]}",
                "speakers": best.get("Speakers") or "",
                "match_score": round(best_score, 4),
                "schedule_title": best["Title"],
                "track": best.get("Stage/Track") or "Unknown",
                "tags": best.get("Tags") or "",
                "day": best.get("Day") or "",
                "time": best.get("Time") or "",
                "format": best.get("Format") or "",
                "schedule_url": best.get("URL") or "",
                "schedule_description": (best.get("Description") or "")[:500],
                "transcript_file": tfile,
            }
            tag = "MANUAL" if num in MANUAL else ("HIGH" if best_score >= 0.5 else "MED" if best_score >= 0.3 else "LOW")
            via = "title" if best_title_sc >= best_d_sc else "descr"
            print(f"  {num} [{tag} {best_score:.2f} via {via}] {yt_title[:55]}", file=sys.stderr)
            print(f"        -> {best['Title'][:70]}", file=sys.stderr)
        new_entries.append(entry)

    print(json.dumps(new_entries, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
